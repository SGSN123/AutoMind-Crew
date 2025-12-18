import openpyxl
import requests
import time
import os
from github import Github, GithubException
from dotenv import load_dotenv

# -------------------------------------------------
# Load environment variables
# -------------------------------------------------
# Works locally (.env) and in GitHub Actions (Secrets)
load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
GITHUB_TOKEN = os.getenv("GH_PAT")
REPO_NAME = os.getenv("GH_REPO_NAME")
REPO_OWNER = os.getenv("GH_USERNAME")
LOCAL_REPO_PATH = os.getenv("LOCAL_REPO_PATH")

# -------------------------------------------------
# Fail fast if env missing
# -------------------------------------------------
missing = []
for key, value in {
    "OPENAI_API_KEY": OPENAI_API_KEY,
    "GH_PAT": GITHUB_TOKEN,
    "GH_REPO_NAME": REPO_NAME,
    "GH_USERNAME": REPO_OWNER
}.items():
    if not value:
        missing.append(key)

if missing:
    raise Exception(f"Missing ENV values: {', '.join(missing)}")

print("✅ Environment variables loaded")

# -------------------------------------------------
# Load Excel (must exist in repo root)
# -------------------------------------------------
wb = openpyxl.load_workbook("Jira.xlsx")
sheet = wb.active

# -------------------------------------------------
# Generate BDD Scenario
# -------------------------------------------------
def generate_bdd_scenario(criteria, scenario_number):
    prompt = f"""
Generate automation-ready Gherkin BDD for Playwright (Chrome only).

Acceptance Criteria:
{criteria}

Rules:
- Valid Gherkin only
- Start with @Scenario_{scenario_number}
- Include Feature
- Use Given When Then And
- One action per step
- Include @smoke and @regression
- No markdown
"""

    response = requests.post(
        "https://api.openai.com/v1/responses",
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json"
        },
        json={
            "model": "gpt-4.1-mini",
            "input": prompt
        }
    )

    response.raise_for_status()
    return response.json()["output"][0]["content"][0]["text"].strip()

# -------------------------------------------------
# Add Background (Chrome only)
# -------------------------------------------------
def add_background(scenario):
    background = """
Background: Login to Appian
  Given I launch Chrome browser
  When I navigate to Appian application
  And I login with valid credentials
"""
    idx = scenario.find("Feature:")
    if idx != -1:
        insert_at = scenario.find("\n", idx) + 1
        return scenario[:insert_at] + background + scenario[insert_at:]
    return background + scenario

# -------------------------------------------------
# Generate JavaScript Step Definitions
# -------------------------------------------------
def generate_steps_js(bdd):
    prompt = f"""
Convert this Gherkin into Playwright Cucumber JavaScript steps.

Rules:
- JavaScript only
- @cucumber/cucumber
- Playwright
- async/await
- Chrome only
- Exact step match
- No comments

BDD:
{bdd}
"""

    response = requests.post(
        "https://api.openai.com/v1/responses",
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json"
        },
        json={
            "model": "gpt-4.1-mini",
            "input": prompt
        }
    )

    response.raise_for_status()
    return response.json()["output"][0]["content"][0]["text"].strip()

# -------------------------------------------------
# GitHub Push Utility
# -------------------------------------------------
def push_file(path, content, message):
    repo = Github(GITHUB_TOKEN).get_repo(f"{REPO_OWNER}/{REPO_NAME}")

    try:
        existing = repo.get_contents(path, ref="main")
        repo.update_file(path, message, content, existing.sha, branch="main")
        print(f"Updated {path}")
    except GithubException as e:
        if e.status == 404:
            repo.create_file(path, message, content, branch="main")
            print(f"Created {path}")
        else:
            raise

    # Optional local copy (for local runs only)
    if LOCAL_REPO_PATH:
        local_path = os.path.join(LOCAL_REPO_PATH, path)
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, "w", encoding="utf-8") as f:
            f.write(content)

# -------------------------------------------------
# Process Excel Rows
# -------------------------------------------------
for row in range(2, sheet.max_row + 1):
    criteria = sheet.cell(row=row, column=12).value  # Column L
    if not criteria:
        continue

    print(f"Processing row {row}")

    bdd = generate_bdd_scenario(criteria, row)
    bdd = add_background(bdd)
    steps = generate_steps_js(bdd)

    sheet.cell(row=row, column=13, value=bdd)

    push_file(
        f"src/test/resources/features/scenario_{row}.feature",
        bdd,
        f"BDD scenario {row}"
    )

    push_file(
        f"src/test/step-definitions/scenario_{row}.steps.js",
        steps,
        f"Steps {row}"
    )

    time.sleep(1)

# -------------------------------------------------
# Save Excel
# -------------------------------------------------
wb.save("updated_acceptance_criteria.xlsx")
print("✅ Excel updated")
print("✅ Features & Steps pushed to GitHub")
